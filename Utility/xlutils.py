import openpyxl
from openpyxl.styles import PatternFill

def getrowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getCoulmnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readdata(file,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnnum).value


def writedata(file,sheetName,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnnum).value=data
    workbook.save(file)


def fillGrennColor(file,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenFill
    workbook.save(file)

def fillredColor(file,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    redfill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum,columnnum).fill=redfill
    workbook.save(file)




